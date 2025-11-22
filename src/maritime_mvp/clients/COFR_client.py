# src/maritime_mvp/clients/COFR_client.py

import io
import os
import re
from datetime import datetime, date
from typing import List, Dict, Any
from urllib.parse import urljoin

import httpx
import openpyxl
import psycopg2
from psycopg2.extras import execute_batch, Json

COFR_STATUS_URL = (
    "https://www.uscg.mil/Mariners/National-Pollution-Funds-Center/COFRs/"
    "ECOFR-Active-Vessel-Status/"
)

# Example DSN: "postgres://user:pass@host:5432/postgres"
PG_DSN = os.environ["DATABASE_URL"]


def _build_client() -> httpx.Client:
    # Use a boring browser UA instead of loudly announcing "hi I'm an ETL bot"
    return httpx.Client(
        timeout=30,
        follow_redirects=True,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
                "(KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36"
            ),
            "Accept": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
                "application/octet-stream;q=0.9,*/*;q=0.8"
            ),
        },
    )


def _discover_xlsx_url(html: str) -> str:
    """
    Scrape the ECOFR Active Vessel Status page to find the XLSX link.
    We look for the first .xlsx href and resolve it relative to COFR_STATUS_URL.
    """
    m = re.search(r'href="([^"]+\.xlsx[^"]*)"', html, re.IGNORECASE)
    if not m:
        raise RuntimeError("Could not find XLSX link on ECOFR Active Vessel Status page")
    href = m.group(1)
    return urljoin(COFR_STATUS_URL, href)


def fetch_xlsx_bytes() -> bytes:
    """Download the current Active Vessel COFR XLSX via the status page."""
    with _build_client() as client:
        # 1) Get the status page
        status_resp = client.get(COFR_STATUS_URL)
        try:
            status_resp.raise_for_status()
        except httpx.HTTPStatusError as exc:
            raise RuntimeError(
                f"Failed to load ECOFR status page: {exc.response.status_code}"
            ) from exc

        # 2) Extract current XLSX link (with ?ver=... token)
        xlsx_url = _discover_xlsx_url(status_resp.text)

        # 3) Download the spreadsheet
        xlsx_resp = client.get(xlsx_url)
        try:
            xlsx_resp.raise_for_status()
        except httpx.HTTPStatusError as exc:
            # Give yourself some debugging context in logs
            snippet = xlsx_resp.text[:200] if xlsx_resp.text else ""
            raise RuntimeError(
                f"Failed to download COFR XLSX ({xlsx_url}): "
                f"{xlsx_resp.status_code} {snippet!r}"
            ) from exc

        return xlsx_resp.content


def parse_rows(data: bytes) -> List[Dict[str, Any]]:
    """Parse XLSX into a list of dict records using real headers."""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    records: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rec: Dict[str, Any] = {}
        for h, v in zip(headers, row):
            if not h:
                continue
            rec[h] = v
        records.append(rec)
    return records


def _to_date(x) -> date | None:
    """Handle Excel datetime, string dates, or None."""
    if x is None:
        return None
    if isinstance(x, datetime):
        return x.date()
    s = str(x).strip()
    if not s:
        return None

    # Try a couple of sane formats
    for fmt in ("%m/%d/%Y %H:%M:%S %p", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Give up gracefully
    return None


def normalize_record(raw: Dict[str, Any]) -> Dict[str, Any]:
    """
    Map Excel columns → DB schema.

    Expected columns:
      Vessel Name, Vin, Vessel Type Code, Vsl Vessel Type Desc,
      Gross Tonnage, Case Control Id, Case- Examiner Id, Case- Operator Name,
      Effective Date, Expiration Date, Insurance Cancel Flag
    """
    vessel_name = (raw.get("Vessel Name") or "").strip()
    vin = (raw.get("Vin") or "").strip()

    return {
        "vessel_name": vessel_name or None,
        "vin": vin or None,
        "vessel_type_code": str(raw.get("Vessel Type Code") or "").strip() or None,
        "vessel_type_desc": str(raw.get("Vsl Vessel Type Desc") or "").strip() or None,
        "gross_tonnage": raw.get("Gross Tonnage"),
        "case_control_id": str(raw.get("Case Control Id") or "").strip() or None,
        "case_examiner_id": str(raw.get("Case- Examiner Id") or "").strip() or None,
        "case_operator_name": str(raw.get("Case- Operator Name") or "").strip() or None,
        "effective_date": _to_date(raw.get("Effective Date")),
        "expiration_date": _to_date(raw.get("Expiration Date")),
        "insurance_cancel_flag": str(raw.get("Insurance Cancel Flag") or "").strip() or None,
        "raw_record": raw,
    }


def upload_to_supabase(records: List[Dict[str, Any]]):
    """Upsert normalized COFR records into Supabase."""
    conn = psycopg2.connect(PG_DSN)
    try:
        with conn, conn.cursor() as cur:
            sql = """
            INSERT INTO public.cofr_active_vessels (
                vessel_name,
                vin,
                vessel_type_code,
                vessel_type_desc,
                gross_tonnage,
                case_control_id,
                case_examiner_id,
                case_operator_name,
                effective_date,
                expiration_date,
                insurance_cancel_flag,
                raw_record,
                last_seen_at
            )
            VALUES (
                %(vessel_name)s,
                %(vin)s,
                %(vessel_type_code)s,
                %(vessel_type_desc)s,
                %(gross_tonnage)s,
                %(case_control_id)s,
                %(case_examiner_id)s,
                %(case_operator_name)s,
                %(effective_date)s,
                %(expiration_date)s,
                %(insurance_cancel_flag)s,
                %(raw_record)s,
                now()
            )
            ON CONFLICT (COALESCE(vin, ''), lower(vessel_name))
            DO UPDATE SET
                vessel_type_code     = EXCLUDED.vessel_type_code,
                vessel_type_desc     = EXCLUDED.vessel_type_desc,
                gross_tonnage        = EXCLUDED.gross_tonnage,
                case_control_id      = EXCLUDED.case_control_id,
                case_examiner_id     = EXCLUDED.case_examiner_id,
                case_operator_name   = EXCLUDED.case_operator_name,
                effective_date       = EXCLUDED.effective_date,
                expiration_date      = EXCLUDED.expiration_date,
                insurance_cancel_flag= EXCLUDED.insurance_cancel_flag,
                raw_record           = EXCLUDED.raw_record,
                last_seen_at         = now();
            """
            batch = []
            for r in records:
                r = r.copy()
                r["raw_record"] = Json(r["raw_record"])
                batch.append(r)
            execute_batch(cur, sql, batch, page_size=500)
    finally:
        conn.close()


def main():
    data = fetch_xlsx_bytes()
    raw_records = parse_rows(data)
    normalized = [normalize_record(r) for r in raw_records if r]
    upload_to_supabase(normalized)


if __name__ == "__main__":
    main()
