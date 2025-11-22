# src/maritime_mvp/connectors/live_sources.py
from __future__ import annotations
import re
import time
import logging
from dataclasses import dataclass, asdict
from decimal import Decimal
from typing import Any, Dict, Optional, Tuple, List
from datetime import datetime
from sqlalchemy import text
import httpx
from lxml import html
import io

# Import the fixed PSIX client
from ..clients.psix_client import PsixClient
from ..db import SessionLocal

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except Exception:
    _OPENPYXL_AVAILABLE = False

logger = logging.getLogger(__name__)

DEFAULT_TIMEOUT = 20
UA = "MaritimeMVP/0.2 (+https://maritime-mvp.onrender.com)"
CACHE_TTL_S = 900  # 15 min default

# Simple in-process TTL cache
_cache: Dict[str, Tuple[float, Any]] = {}

def _get_cached(key: str) -> Optional[Any]:
    v = _cache.get(key)
    if not v:
        return None
    exp, data = v
    if exp < time.time():
        _cache.pop(key, None)
        return None
    return data

def _set_cached(key: str, value: Any, ttl_s: int = CACHE_TTL_S) -> None:
    _cache[key] = (time.time() + ttl_s, value)

@dataclass
class VesselDoc:
    name: str
    expires_on: Optional[str]
    status: Optional[str]
    source: str

@dataclass
class PilotageInfo:
    provider: str
    url: str
    title: Optional[str]
    draft_limits: Optional[Dict[str, Any]]
    boarding_grounds: Optional[str]
    vhf_channel: Optional[str]
    advance_notice: Optional[str]

@dataclass
class LiveBundle:
    vessel: Dict[str, Any]              
    documents: List[VesselDoc]          
    pilotage: Dict[str, Any]            
    marine_exchange: Dict[str, Any]     
    misp: Dict[str, Any]                
    cofr: Dict[str, Any]                
    alerts: List[str]                   

# ---- generic HTML helpers ----------------------------------------------------

def fetch_html(url: str, *, ttl: int = CACHE_TTL_S, parse_extra: bool = False) -> Dict[str, Any]:
    """Fetch and lightly parse HTML pages for relevant maritime info."""
    ck = f"html::{url}"
    cached = _get_cached(ck)
    if cached:
        return cached
    
    try:
        with httpx.Client(timeout=DEFAULT_TIMEOUT, headers={"User-Agent": UA}, verify=False) as client:
            r = client.get(url, follow_redirects=True)
            r.raise_for_status()
            
            try:
                tree = html.fromstring(r.text)
                title = (tree.xpath("//title/text()") or [""])[0].strip()
                
                # Extract visible text
                text_bits = tree.xpath("//body//*[not(self::script or self::style)]/text()")
                text = " ".join(t.strip() for t in text_bits if t.strip())
                
                # Look for common maritime patterns
                extra = {}
                if parse_extra:
                    # VHF channels
                    vhf_match = re.search(r"VHF.*?Channel[s]?\s*(\d+[A-B]?)", text, re.IGNORECASE)
                    if vhf_match:
                        extra["vhf_channel"] = vhf_match.group(1)
                    
                    # Draft limits
                    draft_match = re.search(r"(?:maximum|max).*?draft.*?(\d+\.?\d*)\s*(?:feet|ft|meters|m)", 
                                          text, re.IGNORECASE)
                    if draft_match:
                        extra["max_draft"] = draft_match.group(1)
                    
                    # Advance notice requirements
                    notice_match = re.search(r"(\d+)\s*(?:hours?|hrs?).*?advance.*?notice", 
                                           text, re.IGNORECASE)
                    if notice_match:
                        extra["advance_notice_hours"] = notice_match.group(1)
                    
                    # Find PDF links (often contain tariffs)
                    pdf_links = tree.xpath("//a[contains(@href, '.pdf')]/@href")
                    if pdf_links:
                        # Make PDF links absolute if they're relative
                        from urllib.parse import urljoin
                        extra["pdf_links"] = [urljoin(url, link) for link in pdf_links[:5]]
                
                snap = {
                    "url": url,
                    "title": title,
                    "text_sample": (text[:1500] + "…") if len(text) > 1500 else text,
                    "fetched_at": int(time.time()),
                    **extra
                }
                _set_cached(ck, snap, ttl)
                return snap
                
            except Exception as e:
                logger.warning(f"Failed to parse HTML from {url}: {e}")
                # Return basic info even if parsing fails
                snap = {
                    "url": url,
                    "title": "Parse error",
                    "text_sample": r.text[:500] if r.text else "",
                    "fetched_at": int(time.time()),
                    "error": "HTML parsing failed"
                }
                _set_cached(ck, snap, ttl)
                return snap
                
    except httpx.HTTPStatusError as e:
        logger.warning(f"HTTP error fetching {url}: {e.response.status_code}")
        return {"url": url, "error": f"HTTP {e.response.status_code}", "fetched_at": int(time.time())}
    except Exception as e:
        logger.warning(f"Failed to fetch {url}: {e}")
        return {"url": url, "error": str(e), "fetched_at": int(time.time())}

# ---- PSIX enhanced wrapper with error handling -------------------------------

def psix_summary_by_name(name: str) -> Dict[str, Any]:
    """Get vessel summary from PSIX by name with caching."""
    ck = f"psix::name::{name.lower()}"
    if (v := _get_cached(ck)) is not None:
        return v
    
    try:
        client = PsixClient()
        data = client.search_by_name(name)
        
        # With the new client, data is already a dict
        rows = (data or {}).get("Table") or []
        summary = rows[0] if rows else {}
        
        # Enhance with computed fields
        if summary:
            summary["_psix_match_count"] = len(rows)
            summary["_psix_fetched_at"] = int(time.time())
            
        _set_cached(ck, summary, 600)
        return summary
        
    except Exception as e:
        logger.error(f"PSIX search failed for {name}: {e}")
        # Cache the failure to avoid hammering a broken service
        empty_result = {"error": str(e), "vessel_name": name}
        _set_cached(ck, empty_result, 60)  # Cache failures for 1 minute
        return empty_result

def psix_summary_by_id(vessel_id: int) -> Dict[str, Any]:
    """Get vessel summary from PSIX by ID with caching."""
    ck = f"psix::id::{vessel_id}"
    if (v := _get_cached(ck)) is not None:
        return v
    
    try:
        client = PsixClient()
        data = client.get_vessel_summary(vessel_id=vessel_id)
        
        # With the new client, data is already a dict
        rows = (data or {}).get("Table") or []
        summary = rows[0] if rows else {}
        
        if summary:
            summary["_psix_fetched_at"] = int(time.time())
            
        _set_cached(ck, summary, 600)
        return summary
        
    except Exception as e:
        logger.error(f"PSIX lookup failed for ID {vessel_id}: {e}")
        empty_result = {"error": str(e), "vessel_id": vessel_id}
        _set_cached(ck, empty_result, 60)
        return empty_result
        
def _doc_field(row: Dict[str, Any], *keys: str) -> Optional[str]:
    """Extract field from PSIX row with case-insensitive key matching."""
    for k in keys:
        val = row.get(k) or row.get(k.lower())
        if val:
            return str(val)
    return None

def extract_docs_from_psix_row(row: Dict[str, Any]) -> List[VesselDoc]:
    """Extract document information from PSIX vessel data."""
    docs: List[VesselDoc] = []
    
    if not row or "error" in row:
        return docs
    
    # Certificate of Documentation
    doc_exp = _doc_field(row, "DocumentationExpirationDate", "documentationexpirationdate")
    doc_stat = _doc_field(row, "DocumentationStatus", "documentationstatus")
    if doc_exp or doc_stat:
        docs.append(VesselDoc(
            name="USCG Certificate of Documentation",
            expires_on=doc_exp,
            status=doc_stat,
            source="PSIX",
        ))
    
    # Certificate of Inspection
    coi_exp = _doc_field(row, "COIExpirationDate", "coiexpirationdate")
    if coi_exp:
        docs.append(VesselDoc(
            name="USCG Certificate of Inspection",
            expires_on=coi_exp,
            status=None,
            source="PSIX",
        ))
    
    # Safety Management Certificate
    smc_exp = _doc_field(row, "SMCExpirationDate", "smcexpirationdate")
    if smc_exp:
        docs.append(VesselDoc(
            name="Safety Management Certificate",
            expires_on=smc_exp,
            status=None,
            source="PSIX",
        ))
    
    return docs

# ---- Regional Pilotage & Marine Exchange Registry ---------------------------

REGISTRY = {
    # Bay Area
    "sf_pilots": {
        "url": "https://sfbarpilots.com/new-operational/",
        "provider": "San Francisco Bar Pilots",
        "vhf": "10",
        "boarding": "SF Pilot Station"
    },
    "sf_mx": {
        "url": "https://www.sfmx.org/bay-area-committees/hsc/",
        "provider": "San Francisco Marine Exchange"
    },
    
    # Southern California
    "la_pilot": {
        "url": "https://www.portoflosangeles.org/business/pilot-service",
        "provider": "Los Angeles Pilot Service",
        "vhf": "73",
        "boarding": "LA/LB Pilot Station"
    },
    "lb_pilot": {
        "url": "https://www.jacobsenpilot.com/pilotage/",
        "provider": "Jacobsen Pilot Service",
        "vhf": "73",
        "boarding": "LA/LB Pilot Station"
    },
    "socal_mx": {
        "url": "https://mxsocal.org/",
        "provider": "Marine Exchange of Southern California"
    },
    
    # Puget Sound
    "ps_pilots": {
        "url": "https://www.pspilots.org/dispatch-information/general-guidelines-for-vessels/",
        "provider": "Puget Sound Pilots",
        "vhf": "13",
        "boarding": "Port Angeles Pilot Station"
    },
    "ps_mx": {
        "url": "https://marexps.com/",
        "provider": "Marine Exchange of Puget Sound"
    },
    
    # Columbia River
    "cr_pilots": {
        "url": "https://colrip.com/",
        "provider": "Columbia River Pilots",
        "vhf": "16/13",
        "boarding": "Astoria Pilot Station"
    },
    "cr_mx": {
        "url": "https://www.pdxmex.com/resources/",
        "provider": "Columbia River Marine Exchange"
    },
    
    # Other California
    "oak_pilot": {
        "url": "https://sfbarpilots.com/new-operational/",
        "provider": "San Francisco Bar Pilots (Oakland)",
        "vhf": "10"
    },
    "stockton_pilot": {
        "url": "https://sfbarpilots.com/new-operational/",
        "provider": "San Francisco Bar Pilots (Stockton)",
        "vhf": "10"
    },
    "sd_pilot": {
        "url": "https://www.sdmaritime.com/pilotage/",
        "provider": "San Diego Harbor Pilots",
        "vhf": "14"
    }
}

def pilot_snapshot_for_region(region: str) -> Dict[str, Any]:
    """Fetch pilotage information for a specific region."""
    pilots = {}
    
    if region == "bay_area":
        keys = ["sf_pilots", "oak_pilot", "stockton_pilot"]
    elif region == "socal":
        keys = ["la_pilot", "lb_pilot", "sd_pilot"]
    elif region == "puget":
        keys = ["ps_pilots"]
    elif region == "columbia":
        keys = ["cr_pilots"]
    else:
        return {}
    
    for key in keys:
        if key in REGISTRY:
            try:
                info = REGISTRY[key]
                snap = fetch_html(info["url"], parse_extra=True)
                snap.update({
                    "provider": info["provider"],
                    "vhf_channel": info.get("vhf"),
                    "boarding_grounds": info.get("boarding")
                })
                pilots[key] = snap
            except Exception as e:
                logger.warning(f"Failed to fetch pilot info for {key}: {e}")
                pilots[key] = {"error": str(e), "provider": REGISTRY[key]["provider"]}
    
    return pilots

def mx_snapshot_for_region(region: str) -> Dict[str, Any]:
    """Fetch Marine Exchange information for a specific region."""
    mx_map = {
        "bay_area": "sf_mx",
        "socal": "socal_mx",
        "puget": "ps_mx",
        "columbia": "cr_mx"
    }
    
    key = mx_map.get(region)
    if key and key in REGISTRY:
        try:
            info = REGISTRY[key]
            snap = fetch_html(info["url"], parse_extra=True)
            snap["provider"] = info["provider"]
            return {"primary": snap}
        except Exception as e:
            logger.warning(f"Failed to fetch MX info for {key}: {e}")
            return {"primary": {"error": str(e), "provider": REGISTRY[key]["provider"]}}
    return {}

# ---- California MISP (Marine Invasive Species Program) ----------------------

MISP_INFO = {
    "program": "California Marine Invasive Species Program (MISP)",
    "sites": [
        "https://www.slc.ca.gov/misp/",
        "https://cdtfa.ca.gov/taxes-and-fees/marine-invasive-species-fee/",
    ],
    "current_fee": "$1000 per voyage (300+ GT vessels)",
    "exemptions": ["Military", "Law enforcement", "Research vessels"]
}

MONEY_RE = re.compile(r"\$[\d,]+(?:\.\d{2})?")

def fetch_misp_snapshot() -> Dict[str, Any]:
    """Fetch current MISP fee information from California sources."""
    snaps = []
    for url in MISP_INFO["sites"]:
        try:
            snap = fetch_html(url)
            snaps.append(snap)
        except Exception as e:
            logger.warning(f"Failed to fetch MISP from {url}: {e}")
            snaps.append({"url": url, "error": str(e)})
    
    # Extract dollar amounts from CDTFA page
    dollars = []
    if len(snaps) > 1 and "text_sample" in snaps[-1]:
        dollars = MONEY_RE.findall(snaps[-1].get("text_sample", ""))
    
    return {
        "program": MISP_INFO["program"],
        "current_fee": MISP_INFO["current_fee"],
        "exemptions": MISP_INFO["exemptions"],
        "pages": snaps,
        "possible_amounts_seen": list(dict.fromkeys(dollars))[:6],  # de-dupe & cap
        "effective_date": "2020-01-01"
    }

# ---- APHIS AQI Commercial Vessel Fees --------------------------------------

APHIS_FEES_URL = "https://www.aphis.usda.gov/aqi/fees"
APHIS_COMM_VESSEL_URL = "https://www.aphis.usda.gov/aqi/commercial-vessel-fee"


def _parse_money_first(text: str, pattern: str) -> Optional[Decimal]:
    m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
    if not m:
        return None
    # choose the first captured dollar string
    for group in m.groups():
        if not group:
            continue
        cleaned = "".join(ch for ch in group if (ch.isdigit() or ch == "."))
        if cleaned:
            try:
                return Decimal(cleaned)
            except Exception:
                continue
    return None


def fetch_aphis_vessel_fees() -> Dict[str, Any]:
    """
    Fetch current APHIS AQI commercial vessel user fees from USDA public pages.

    Returns:
      {
        "standard_fee": Decimal or None,  # Commercial Vessel
        "cascadia_fee": Decimal or None, # Commercial Vessel - Great Lakes/Cascadia
        "sources": [...],                # basic page metadata
      }
    """

    ck = "aphis::vessel_fees"
    cached = _get_cached(ck)
    if cached is not None:
        return cached

    sources: List[Dict[str, Any]] = []

    # 1) Standard commercial vessel fee (dedicated page)
    std_val: Optional[Decimal] = None
    try:
        std_snap = fetch_html(APHIS_COMM_VESSEL_URL)
        sources.append(std_snap)
        txt = std_snap.get("text_sample", "") or ""
        std_val = _parse_money_first(txt, r"current commercial vessel fee is\s*\$([\d,]+\.\d+)")
    except Exception as e:
        logger.warning(f"Failed to fetch APHIS commercial vessel fee page: {e}")

    # 2) Great Lakes / Cascadia reduced fee from AQI fees table
    cas_val: Optional[Decimal] = None
    try:
        tbl_snap = fetch_html(APHIS_FEES_URL)
        sources.append(tbl_snap)
        txt = tbl_snap.get("text_sample", "") or ""
        # Row: "Commercial Vessel- Great Lakes/Cascadia ... $prev $current ..."
        cas_val = _parse_money_first(
            txt,
            r"Commercial Vessel-?\s*Great Lakes/Cascadia.*?\$([\d,]+\.\d+)\s*\$([\d,]+\.\d+)"
        )
    except Exception as e:
        logger.warning(f"Failed to fetch APHIS AQI fees table: {e}")

    result = {
        "standard_fee": std_val,
        "cascadia_fee": cas_val,
        "sources": sources,
    }
    _set_cached(ck, result, ttl_s=3600)  # 1 hour cache
    return result

# ---- COFR (Certificate of Financial Responsibility) -------------------------

COFR_URLS = {
    "search": "https://publicsearch.npfc.uscg.mil/COFR/Default.aspx",
    "active_list": "https://www.uscg.mil/Mariners/National-Pollution-Funds-Center/COFRs/ECOFR-Active-Vessel-Status/",
    "api_check": "https://cgmix.uscg.mil/xml/COFRData.asmx?WSDL"  # if exists
}

COFR_ACTIVE_XLSX = (
    "https://www.uscg.mil/Portals/0/NPFC/COFR/"
    "ECOFR%20active%20vessel%20cofr.xlsx"
)


def _fetch_cofr_from_db(
    imo_or_official_no: Optional[str],
    vessel_name: Optional[str]
) -> Optional[Dict[str, Any]]:
    """
    Look up an active COFR record in cofr_active_vessels by VIN or vessel name.
    Note: this file doesn't provide IMO, only Vin.
    """
    norm_id = (imo_or_official_no or "").strip()
    norm_name = re.sub(r"\W+", "", (vessel_name or "").upper())

    with SessionLocal() as db:
        # 1) Try exact VIN match
        if norm_id:
            row = db.execute(
                text(
                    """
                    SELECT vessel_name,
                           vin,
                           vessel_type_code,
                           vessel_type_desc,
                           gross_tonnage,
                           status,
                           expiration_date,
                           insurance_cancel_flag,
                           raw_expiry
                    FROM (
                        SELECT
                          vessel_name,
                          vin,
                          vessel_type_code,
                          vessel_type_desc,
                          gross_tonnage,
                          insurance_cancel_flag AS status,
                          expiration_date,
                          raw_record->>'Expiration Date' AS raw_expiry
                        FROM cofr_active_vessels
                    ) t
                    WHERE vin = :id
                    LIMIT 1
                """
                ),
                {"id": norm_id},
            ).fetchone()
            if row:
                return {
                    "vessel_name": row[0],
                    "vin": row[1],
                    "vessel_type_code": row[2],
                    "vessel_type_desc": row[3],
                    "gross_tonnage": row[4],
                    "status": row[5],
                    "expiry_date": row[6].isoformat() if row[6] else None,
                    "raw_expiry": row[7],
                }

        # 2) Fallback to name match
        if norm_name:
            row = db.execute(
                text(
                    """
                    SELECT vessel_name,
                           vin,
                           vessel_type_code,
                           vessel_type_desc,
                           gross_tonnage,
                           insurance_cancel_flag AS status,
                           expiration_date,
                           raw_record->>'Expiration Date' AS raw_expiry
                    FROM cofr_active_vessels
                    WHERE regexp_replace(upper(vessel_name), '\\W+', '', 'g') = :nm
                    LIMIT 1
                """
                ),
                {"nm": norm_name},
            ).fetchone()
            if row:
                return {
                    "vessel_name": row[0],
                    "vin": row[1],
                    "vessel_type_code": row[2],
                    "vessel_type_desc": row[3],
                    "gross_tonnage": row[4],
                    "status": row[5],
                    "expiry_date": row[6].isoformat() if row[6] else None,
                    "raw_expiry": row[7],
                }

    return None


def _fetch_cofr_active_rows(ttl_s: int = 3600) -> List[Dict[str, Any]]:
    """
    Download and parse the ECOFR active vessel spreadsheet into a list of dicts.

    Each dict is {header: value, ...}, with headers taken from the first row.
    """
    ck = "cofr::active_rows"
    cached = _get_cached(ck)
    if cached is not None:
        return cached

    if not _OPENPYXL_AVAILABLE:
        logger.warning("openpyxl not available; skipping COFR active list parsing")
        _set_cached(ck, [])
        return []

    try:
        with httpx.Client(timeout=DEFAULT_TIMEOUT, headers={"User-Agent": UA}, verify=False) as client:
            r = client.get(COFR_ACTIVE_XLSX, follow_redirects=True)
            r.raise_for_status()
            data = r.content
    except Exception as e:
        logger.warning(f"Failed to fetch COFR active list: {e}")
        _set_cached(ck, [])
        return []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h is not None else "" for h in header_row]

        rows: List[Dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            rec: Dict[str, Any] = {}
            for h, v in zip(headers, row):
                if not h:
                    continue
                rec[h] = v
            rows.append(rec)

        _set_cached(ck, rows, ttl_s)
        return rows
    except Exception as e:
        logger.warning(f"Failed to parse COFR active list: {e}")
        _set_cached(ck, [])
        return []


def _cofr_get_field(rec: Dict[str, Any], *predicates) -> str:
    """
    Find first column whose header matches one of the given predicates.
    predicates are callables that take header:str -> bool.
    """
    for h, v in rec.items():
        h_lower = str(h).lower()
        for pred in predicates:
            if pred(h_lower):
                return "" if v is None else str(v).strip()
    return ""


def _match_cofr_record(
    rows: List[Dict[str, Any]],
    vessel_name: Optional[str],
    imo_or_official_no: Optional[str],
) -> Optional[Dict[str, Any]]:
    if not rows:
        return None

    norm_id = (imo_or_official_no or "").strip().lstrip("0")
    norm_name = re.sub(r"\W+", "", (vessel_name or "").upper())

    best: Optional[Dict[str, Any]] = None

    # 1) Strong match on IMO / Official #
    if norm_id:
        for rec in rows:
            id_candidate = _cofr_get_field(
                rec,
                lambda h: "imo" in h,
                lambda h: "official" in h and "number" in h,
            )
            cand_norm = id_candidate.strip().lstrip("0")
            if cand_norm and cand_norm == norm_id:
                best = rec
                break

    # 2) Fallback to exact vessel name match (sanitized)
    if best is None and norm_name:
        for rec in rows:
            name_candidate = _cofr_get_field(
                rec,
                lambda h: "vessel" in h and "name" in h,
                lambda h: h == "name",
            )
            cand_norm = re.sub(r"\W+", "", name_candidate.upper())
            if cand_norm and cand_norm == norm_name:
                best = rec
                break

    if best is None:
        return None

    raw_expiry = _cofr_get_field(
        best,
        lambda h: "exp" in h and "date" in h,
        lambda h: "valid" in h and "thru" in h,
    )
    status = _cofr_get_field(
        best,
        lambda h: "status" in h,
        lambda h: "cofr" in h and "status" in h,
    )
    vessel_nm = _cofr_get_field(
        best,
        lambda h: "vessel" in h and "name" in h,
        lambda h: h == "name",
    )
    imo_val = _cofr_get_field(best, lambda h: "imo" in h)
    off_val = _cofr_get_field(best, lambda h: "official" in h and "number" in h)

    # Normalize expiry into ISO if possible
    exp_iso: Optional[str] = None
    if raw_expiry:
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y"):
            try:
                exp_iso = datetime.strptime(raw_expiry.strip()[:10], fmt).date().isoformat()
                break
            except Exception:
                continue

    return {
        "vessel_name": vessel_nm or vessel_name,
        "imo_number": imo_val or None,
        "official_number": off_val or None,
        "raw_expiry": raw_expiry or None,
        "expiry_date": exp_iso,
        "status": status or None,
    }

def cofr_snapshot(vessel_name: Optional[str], imo_or_official_no: Optional[str]) -> Dict[str, Any]:
    """Get COFR lookup information and, when possible, an active record with expiry."""
    # 1) Try DB first
    active_record: Optional[Dict[str, Any]] = None
    try:
        active_record = _fetch_cofr_from_db(imo_or_official_no, vessel_name)
    except Exception as e:
        logger.warning(f"COFR DB lookup failed: {e}")

    # 2) (Optional) still fetch the HTML page for guidance only
    try:
        snap = fetch_html(COFR_URLS["search"], parse_extra=True)
    except Exception as e:
        logger.warning(f"Failed to fetch COFR page: {e}")
        snap = {"url": COFR_URLS["search"], "error": str(e)}

    guidance = []
    if vessel_name:
        guidance.append(f"Search by vessel name: '{vessel_name}'")
    if imo_or_official_no:
        guidance.append(f"Search by IMO/Official #: '{imo_or_official_no}'")

    cofr_required = {
        "tankers_over_300gt": True,
        "vessels_over_400gt": True,
        "exceptions": ["Public vessels", "Oil spill response vessels"],
    }

    return {
        "entrypoint": COFR_URLS["search"],
        "active_list": COFR_URLS["active_list"],
        "page": snap,
        "query": {"vessel_name": vessel_name, "id": imo_or_official_no},
        "search_guidance": guidance,
        "requirements": cofr_required,
        "active_record": active_record,
    }

# ---- Additional Document Checks ----------------------------------------------

def check_document_alerts(docs: List[VesselDoc]) -> List[str]:
    """Generate alerts for missing or expiring documents."""
    alerts = []
    
    # Check for expiring docs
    from datetime import datetime, timedelta
    today = datetime.now().date()
    warning_days = 30
    
    for doc in docs:
        if doc.expires_on:
            try:
                # Handle various date formats
                date_str = doc.expires_on[:10] if len(doc.expires_on) >= 10 else doc.expires_on
                exp_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                days_until = (exp_date - today).days
                
                if days_until < 0:
                    alerts.append(f"⚠️ {doc.name} EXPIRED {abs(days_until)} days ago")
                elif days_until <= warning_days:
                    alerts.append(f"⚠️ {doc.name} expires in {days_until} days")
            except (ValueError, TypeError) as e:
                logger.warning(f"Could not parse date {doc.expires_on}: {e}")
    
    # Check for missing critical docs
    doc_names = [d.name.lower() for d in docs]
    doc_names_str = " ".join(doc_names)
    
    if not docs or "certificate of documentation" not in doc_names_str:
        alerts.append("📋 Certificate of Documentation not found in PSIX")
    
    return alerts

# ---- Region Detection --------------------------------------------------------

def choose_region(port_code: Optional[str], port_name: Optional[str], 
                 state: Optional[str], is_cascadia: Optional[bool]) -> str:
    """Determine the maritime region based on port information."""
    
    # Port code mapping (your internal codes)
    code_map = {
        "LALB": "socal",
        "SFBAY": "bay_area",
        "PUGET": "puget",
        "COLRIV": "columbia",
        "STKN": "bay_area",
        "OAK": "bay_area",
        "SDG": "socal",
        "HUM": "bay_area",
        "GRH": "columbia",
        "VAN": "columbia",
        "EVR": "puget"
    }
    
    if port_code and port_code in code_map:
        return code_map[port_code]
    
    # Name-based detection
    name = (port_name or "").lower()
    if any(x in name for x in ["san francisco", "oakland", "richmond", "stockton", "sacramento", "alameda", "redwood"]):
        return "bay_area"
    if any(x in name for x in ["los angeles", "long beach", "san diego", "hueneme", "port hueneme"]):
        return "socal"
    if any(x in name for x in ["seattle", "tacoma", "everett", "olympia", "bellingham", "anacortes"]):
        return "puget"
    if any(x in name for x in ["portland", "astoria", "columbia", "vancouver usa", "longview", "kalama"]):
        return "columbia"
    
    # State-based fallback
    st = (state or "").upper()
    if st == "CA":
        return "bay_area"  # default California
    elif st == "WA" or is_cascadia:
        return "puget"
    elif st == "OR":
        return "columbia"
    
    return "bay_area"  # final default

# ---- Main Orchestrator -------------------------------------------------------

def build_live_bundle(*,
                     vessel_name: Optional[str] = None,
                     vessel_id: Optional[int] = None,
                     port_code: Optional[str] = None,
                     port_name: Optional[str] = None,
                     state: Optional[str] = None,
                     is_cascadia: Optional[bool] = None,
                     imo_or_official_no: Optional[str] = None) -> Dict[str, Any]:
    """
    Build a comprehensive bundle of live maritime data for a vessel and port.
    
    This orchestrates calls to multiple sources and returns a unified response
    with vessel info, documents, pilotage, marine exchange, fees, and alerts.
    """
    
    logger.info(f"Building live bundle for vessel={vessel_name}, port={port_code}")
    
    # 1) Fetch vessel data from PSIX with error handling
    vrow = {}
    try:
        if vessel_id is not None:
            vrow = psix_summary_by_id(vessel_id)
        elif vessel_name:
            vrow = psix_summary_by_name(vessel_name)
        
        # Check for errors in the result
        if vrow and "error" in vrow:
            logger.warning(f"PSIX returned error: {vrow.get('error')}")
            vrow = {}  # Use empty dict if there was an error
    except Exception as e:
        logger.error(f"Exception getting PSIX data: {e}")
        vrow = {}
    
    docs: List[VesselDoc] = extract_docs_from_psix_row(vrow) if vrow else []

    # 2) Region + pilot/MX/MISP
    region = choose_region(port_code, port_name, state, is_cascadia)
    logger.info(f"Selected region: {region}")

    # Fetch regional information with error handling
    pilot = {}
    mx = {}
    misp = {}
    cofr_data = {}
    
    try:
        pilot = pilot_snapshot_for_region(region)
    except Exception as e:
        logger.warning(f"Failed to get pilotage info: {e}")
    
    try:
        mx = mx_snapshot_for_region(region)
    except Exception as e:
        logger.warning(f"Failed to get marine exchange info: {e}")
    
    # California-specific fees
    if (state or "").upper() == "CA":
        try:
            misp = fetch_misp_snapshot()
        except Exception as e:
            logger.warning(f"Failed to get MISP info: {e}")
    
    # 3) COFR (now with active record)
    try:
        cofr_data = cofr_snapshot(
            vessel_name=vessel_name or vrow.get("VesselName") or vrow.get("vesselname"),
            imo_or_official_no=imo_or_official_no or vrow.get("IMONumber") or vrow.get("OfficialNumber")
        )
    except Exception as e:
        logger.warning(f"Failed to get COFR info: {e}")
        cofr_data = {"error": str(e)}

    active = cofr_data.get("active_record") or {}
    if active.get("expiry_date") or active.get("raw_expiry"):
        docs.append(
            VesselDoc(
                name="Certificate of Financial Responsibility (COFR)",
                expires_on=active.get("expiry_date") or active.get("raw_expiry"),
                status=active.get("status") or "Active",
                source="NPFC",
            )
        )

    # 4) Now compute alerts with COFR included
    alerts = check_document_alerts(docs)

    # 5) Build final bundle
    bundle = LiveBundle(
        vessel=vrow,
        documents=[asdict(d) for d in docs],
        pilotage=pilot,
        marine_exchange=mx,
        misp=misp,
        cofr=cofr_data,
        alerts=alerts
    )
    
    return asdict(bundle)

# ---- Utility Functions -------------------------------------------------------

def clear_cache():
    """Clear the in-process cache (useful for testing or forced refresh)."""
    global _cache
    _cache.clear()
    logger.info("Cache cleared")

def get_cache_stats() -> Dict[str, int]:
    """Get current cache statistics."""
    now = time.time()
    active = sum(1 for exp, _ in _cache.values() if exp > now)
    expired = len(_cache) - active
    return {
        "total_entries": len(_cache),
        "active_entries": active,
        "expired_entries": expired
    }
